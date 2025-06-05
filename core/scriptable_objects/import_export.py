"""
Import/Export System for Scriptable Objects
Handles importing and exporting templates and instances
"""

import json
import zipfile
import csv
from typing import Dict, List, Any, Optional, Union
from pathlib import Path
import tempfile
import shutil

from .template import ScriptableObjectTemplate
from .instance import ScriptableObjectInstance

# Use TYPE_CHECKING to avoid circular imports
from typing import TYPE_CHECKING
if TYPE_CHECKING:
    from .manager import ScriptableObjectManager


class ImportExportManager:
    """Manages import and export operations"""
    
    def __init__(self, manager: ScriptableObjectManager):
        self.manager = manager
    
    # Template Export/Import
    
    def export_template(self, template_name: str, file_path: str, format: str = "json") -> bool:
        """Export a single template"""
        template = self.manager.get_template(template_name)
        if not template:
            return False
        
        try:
            if format.lower() == "json":
                self._export_template_json(template, file_path)
            elif format.lower() == "yaml":
                self._export_template_yaml(template, file_path)
            else:
                return False
            return True
        except Exception as e:
            print(f"Error exporting template: {e}")
            return False
    
    def export_all_templates(self, file_path: str, format: str = "zip") -> bool:
        """Export all templates"""
        try:
            if format.lower() == "zip":
                return self._export_templates_zip(file_path)
            elif format.lower() == "json":
                return self._export_templates_json(file_path)
            else:
                return False
        except Exception as e:
            print(f"Error exporting templates: {e}")
            return False
    
    def import_template(self, file_path: str, overwrite: bool = False) -> bool:
        """Import a single template"""
        try:
            path = Path(file_path)
            
            if path.suffix.lower() == ".json":
                return self._import_template_json(file_path, overwrite)
            elif path.suffix.lower() in [".yaml", ".yml"]:
                return self._import_template_yaml(file_path, overwrite)
            else:
                return False
        except Exception as e:
            print(f"Error importing template: {e}")
            return False
    
    def import_templates(self, file_path: str, overwrite: bool = False) -> Dict[str, bool]:
        """Import multiple templates"""
        results = {}
        
        try:
            path = Path(file_path)
            
            if path.suffix.lower() == ".zip":
                results = self._import_templates_zip(file_path, overwrite)
            elif path.suffix.lower() == ".json":
                results = self._import_templates_json(file_path, overwrite)
            else:
                results["error"] = False
        except Exception as e:
            print(f"Error importing templates: {e}")
            results["error"] = False
        
        return results
    
    # Instance Export/Import
    
    def export_instances(self, template_name: str, file_path: str, format: str = "json") -> bool:
        """Export instances of a template"""
        instances = self.manager.get_instances_of_template(template_name)
        if not instances:
            return False
        
        try:
            if format.lower() == "json":
                return self._export_instances_json(instances, file_path)
            elif format.lower() == "csv":
                return self._export_instances_csv(template_name, instances, file_path)
            elif format.lower() == "xlsx":
                return self._export_instances_xlsx(template_name, instances, file_path)
            else:
                return False
        except Exception as e:
            print(f"Error exporting instances: {e}")
            return False
    
    def import_instances(self, template_name: str, file_path: str, overwrite: bool = False) -> Dict[str, bool]:
        """Import instances for a template"""
        results = {}
        
        try:
            path = Path(file_path)
            
            if path.suffix.lower() == ".json":
                results = self._import_instances_json(template_name, file_path, overwrite)
            elif path.suffix.lower() == ".csv":
                results = self._import_instances_csv(template_name, file_path, overwrite)
            elif path.suffix.lower() == ".xlsx":
                results = self._import_instances_xlsx(template_name, file_path, overwrite)
            else:
                results["error"] = False
        except Exception as e:
            print(f"Error importing instances: {e}")
            results["error"] = False
        
        return results
    
    # Private implementation methods
    
    def _export_template_json(self, template: ScriptableObjectTemplate, file_path: str):
        """Export template to JSON"""
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(template.to_dict(), f, indent=2, ensure_ascii=False)
    
    def _export_template_yaml(self, template: ScriptableObjectTemplate, file_path: str):
        """Export template to YAML"""
        try:
            import yaml
            with open(file_path, 'w', encoding='utf-8') as f:
                yaml.dump(template.to_dict(), f, default_flow_style=False, allow_unicode=True)
        except ImportError:
            raise ImportError("PyYAML is required for YAML export")
    
    def _export_templates_zip(self, file_path: str) -> bool:
        """Export all templates to a ZIP file"""
        with zipfile.ZipFile(file_path, 'w', zipfile.ZIP_DEFLATED) as zf:
            for template in self.manager.get_all_templates():
                template_data = json.dumps(template.to_dict(), indent=2, ensure_ascii=False)
                zf.writestr(f"templates/{template.name}.json", template_data)
        return True
    
    def _export_templates_json(self, file_path: str) -> bool:
        """Export all templates to a single JSON file"""
        templates_data = {
            "templates": [template.to_dict() for template in self.manager.get_all_templates()],
            "export_version": "1.0",
            "export_type": "templates"
        }
        
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(templates_data, f, indent=2, ensure_ascii=False)
        return True
    
    def _import_template_json(self, file_path: str, overwrite: bool) -> bool:
        """Import template from JSON"""
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        template = ScriptableObjectTemplate.from_dict(data)
        
        # Check if template already exists
        if self.manager.get_template(template.name) and not overwrite:
            return False
        
        self.manager.save_template(template)
        return True
    
    def _import_template_yaml(self, file_path: str, overwrite: bool) -> bool:
        """Import template from YAML"""
        try:
            import yaml
            with open(file_path, 'r', encoding='utf-8') as f:
                data = yaml.safe_load(f)
            
            template = ScriptableObjectTemplate.from_dict(data)
            
            # Check if template already exists
            if self.manager.get_template(template.name) and not overwrite:
                return False
            
            self.manager.save_template(template)
            return True
        except ImportError:
            raise ImportError("PyYAML is required for YAML import")
    
    def _import_templates_zip(self, file_path: str, overwrite: bool) -> Dict[str, bool]:
        """Import templates from ZIP file"""
        results = {}
        
        with zipfile.ZipFile(file_path, 'r') as zf:
            for file_info in zf.filelist:
                if file_info.filename.startswith("templates/") and file_info.filename.endswith(".json"):
                    template_name = Path(file_info.filename).stem
                    
                    try:
                        with zf.open(file_info.filename) as f:
                            data = json.load(f)
                        
                        template = ScriptableObjectTemplate.from_dict(data)
                        
                        # Check if template already exists
                        if self.manager.get_template(template.name) and not overwrite:
                            results[template_name] = False
                            continue
                        
                        self.manager.save_template(template)
                        results[template_name] = True
                    except Exception as e:
                        print(f"Error importing template {template_name}: {e}")
                        results[template_name] = False
        
        return results
    
    def _import_templates_json(self, file_path: str, overwrite: bool) -> Dict[str, bool]:
        """Import templates from single JSON file"""
        results = {}
        
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        templates_data = data.get("templates", [])
        
        for template_data in templates_data:
            try:
                template = ScriptableObjectTemplate.from_dict(template_data)
                
                # Check if template already exists
                if self.manager.get_template(template.name) and not overwrite:
                    results[template.name] = False
                    continue
                
                self.manager.save_template(template)
                results[template.name] = True
            except Exception as e:
                print(f"Error importing template {template_data.get('name', 'unknown')}: {e}")
                results[template_data.get('name', 'unknown')] = False
        
        return results
    
    def _export_instances_json(self, instances: List[ScriptableObjectInstance], file_path: str) -> bool:
        """Export instances to JSON"""
        instances_data = {
            "instances": [instance.to_dict() for instance in instances],
            "export_version": "1.0",
            "export_type": "instances"
        }
        
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(instances_data, f, indent=2, ensure_ascii=False)
        return True
    
    def _export_instances_csv(self, template_name: str, instances: List[ScriptableObjectInstance], file_path: str) -> bool:
        """Export instances to CSV"""
        if not instances:
            return False
        
        template = self.manager.get_template(template_name)
        if not template:
            return False
        
        # Get all field names
        field_names = ["name", "instance_id"] + [field.name for field in template.fields]
        
        with open(file_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.writer(f)
            
            # Write header
            writer.writerow(field_names)
            
            # Write data
            for instance in instances:
                row = [instance.name, instance.instance_id]
                for field in template.fields:
                    value = instance.get_value(field.name, "")
                    # Convert complex types to JSON strings
                    if isinstance(value, (list, dict)):
                        value = json.dumps(value)
                    row.append(value)
                writer.writerow(row)
        
        return True
    
    def _export_instances_xlsx(self, template_name: str, instances: List[ScriptableObjectInstance], file_path: str) -> bool:
        """Export instances to Excel"""
        try:
            import openpyxl
            from openpyxl import Workbook
            
            if not instances:
                return False
            
            template = self.manager.get_template(template_name)
            if not template:
                return False
            
            wb = Workbook()
            ws = wb.active
            ws.title = template_name
            
            # Write header
            headers = ["Name", "Instance ID"] + [field.name for field in template.fields]
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
            
            # Write data
            for row, instance in enumerate(instances, 2):
                ws.cell(row=row, column=1, value=instance.name)
                ws.cell(row=row, column=2, value=instance.instance_id)
                
                for col, field in enumerate(template.fields, 3):
                    value = instance.get_value(field.name, "")
                    # Convert complex types to JSON strings
                    if isinstance(value, (list, dict)):
                        value = json.dumps(value)
                    ws.cell(row=row, column=col, value=value)
            
            wb.save(file_path)
            return True
        except ImportError:
            raise ImportError("openpyxl is required for Excel export")
    
    def _import_instances_json(self, template_name: str, file_path: str, overwrite: bool) -> Dict[str, bool]:
        """Import instances from JSON"""
        results = {}
        
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        instances_data = data.get("instances", [])
        
        for instance_data in instances_data:
            try:
                instance = ScriptableObjectInstance.from_dict(instance_data)
                
                # Check if instance already exists
                existing = self.manager.get_instance(template_name, instance.instance_id)
                if existing and not overwrite:
                    results[instance.name] = False
                    continue
                
                self.manager.save_instance(instance)
                results[instance.name] = True
            except Exception as e:
                print(f"Error importing instance {instance_data.get('name', 'unknown')}: {e}")
                results[instance_data.get('name', 'unknown')] = False
        
        return results
    
    def _import_instances_csv(self, template_name: str, file_path: str, overwrite: bool) -> Dict[str, bool]:
        """Import instances from CSV"""
        results = {}
        template = self.manager.get_template(template_name)
        if not template:
            return {"error": False}
        
        with open(file_path, 'r', encoding='utf-8') as f:
            reader = csv.DictReader(f)
            
            for row in reader:
                try:
                    instance_name = row.get("name", "")
                    instance_id = row.get("instance_id", "")
                    
                    # Create instance data
                    instance_data = {}
                    for field in template.fields:
                        if field.name in row:
                            value = row[field.name]
                            # Try to parse JSON for complex types
                            if value and field.field_type.value in ["array", "object"]:
                                try:
                                    value = json.loads(value)
                                except json.JSONDecodeError:
                                    pass
                            instance_data[field.name] = value
                    
                    instance = ScriptableObjectInstance(
                        template_name=template_name,
                        instance_id=instance_id,
                        name=instance_name,
                        **instance_data
                    )
                    
                    # Check if instance already exists
                    existing = self.manager.get_instance(template_name, instance.instance_id)
                    if existing and not overwrite:
                        results[instance_name] = False
                        continue
                    
                    self.manager.save_instance(instance)
                    results[instance_name] = True
                except Exception as e:
                    print(f"Error importing instance {row.get('name', 'unknown')}: {e}")
                    results[row.get('name', 'unknown')] = False
        
        return results
    
    def _import_instances_xlsx(self, template_name: str, file_path: str, overwrite: bool) -> Dict[str, bool]:
        """Import instances from Excel"""
        try:
            import openpyxl
            
            results = {}
            template = self.manager.get_template(template_name)
            if not template:
                return {"error": False}
            
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            
            # Get headers from first row
            headers = [cell.value for cell in ws[1]]
            
            # Process data rows
            for row in ws.iter_rows(min_row=2, values_only=True):
                try:
                    row_data = dict(zip(headers, row))
                    
                    instance_name = row_data.get("Name", "")
                    instance_id = row_data.get("Instance ID", "")
                    
                    # Create instance data
                    instance_data = {}
                    for field in template.fields:
                        if field.name in row_data:
                            value = row_data[field.name]
                            # Try to parse JSON for complex types
                            if value and field.field_type.value in ["array", "object"]:
                                try:
                                    value = json.loads(value)
                                except (json.JSONDecodeError, TypeError):
                                    pass
                            instance_data[field.name] = value
                    
                    instance = ScriptableObjectInstance(
                        template_name=template_name,
                        instance_id=instance_id,
                        name=instance_name,
                        **instance_data
                    )
                    
                    # Check if instance already exists
                    existing = self.manager.get_instance(template_name, instance.instance_id)
                    if existing and not overwrite:
                        results[instance_name] = False
                        continue
                    
                    self.manager.save_instance(instance)
                    results[instance_name] = True
                except Exception as e:
                    print(f"Error importing instance {row_data.get('Name', 'unknown')}: {e}")
                    results[row_data.get('Name', 'unknown')] = False
            
            return results
        except ImportError:
            raise ImportError("openpyxl is required for Excel import")
